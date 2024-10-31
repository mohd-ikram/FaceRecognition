from openpyxl import load_workbook 
from openpyxl import Workbook 
from datetime import datetime
import threading


""" Mark attendance of person in excel sheet taking input as name of person.
Param checkin shwoing the peson is checking in or out"""

lock = threading.Lock()
def mark_attendance(name, checkin = True, filename = "record_sheet.xlsx"):
    # accuire lock
    lock.acquire()
    # open the excel workbook
    try:
        wb = load_workbook(filename=filename)
    except FileNotFoundError:
        print(f"File {filename} not found. Creating new file with the same name")
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Name", "Checkin", "Checkout"])
    
    # Get the active worksheet
    sheet = wb.active
    # Header coloms
    header_cols = [cell.value for cell in sheet[1]]
    # Find the row of named person
    name_col = header_cols.index("Name")+1
    checkin_col = header_cols.index("Checkin")+1
    checkout_col = header_cols.index("Checkout")+1
    print(name_col,checkin_col,checkout_col)
    checkin_time = ""
    checkout_time = ""
    if checkin == True: 
        checkin_time = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    else:
        checkout_time = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column= name_col).value == name:
            # sheet.cell(row=row, column = len(header_cols)-1).value = current_date_time
            # sheet.cell(row=row, column = name_col).value = name
            if checkin == True:
                sheet.cell(row=row, column = checkin_col).value = checkin_time
                sheet.cell(row=row, column = checkout_col).value = checkout_time
            else:
                sheet.cell(row=row, column = checkout_col).value = checkout_time
            # print("Modified")
            break
    else:
        sheet.append([name, checkin_time, checkout_time])
        print("Appended")

    wb.save(filename=filename)
    wb.close()
    lock.release()

# mark_attendance("Ikram",True)

